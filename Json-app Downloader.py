#!/usr/bin/env python3
"""
Use the command 
python filename.py --json filename.json --outdir . --threads 4
"""

import os
import sys
import re
import json
import time
import zipfile
import shutil
import logging
import argparse
import tempfile
import subprocess
from pathlib import Path

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/140.0.0.0 Safari/537.36"
}
APKTOOL_FILENAME = "apktool.jar"
APKTOOL_URL = "https://bitbucket.org/iBotPeaches/apktool/downloads/apktool_2.12.1.jar"
TEMP_DIR = ".temp_apks"
MANIFEST_DIR = ".manifests"


def check_install_library(lib_name):
    try:
        __import__(lib_name)
    except ImportError:
        print(f"[!] Library '{lib_name}' not found. Installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", lib_name])
        print(f"[✓] Library '{lib_name}' installed.")

for lib in ["requests", "bs4", "tqdm", "openpyxl"]:
    check_install_library(lib)

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font

def ensure_apktool(script_dir, logger):
    apktool_path = Path(script_dir) / APKTOOL_FILENAME
    if not apktool_path.exists():
        logger.info("[!] apktool.jar not found. Downloading...")
        r = requests.get(APKTOOL_URL, headers=HEADERS, stream=True, timeout=30)
        r.raise_for_status()
        with open(apktool_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        logger.info(f"[✓] Apktool downloaded -> {apktool_path}")
    return apktool_path

def safe_request(session, method, url, logger, retries=5, backoff=2, **kwargs):
    for attempt in range(1, retries + 1):
        try:
            if method.lower() == "get":
                r = session.get(url, headers=HEADERS, timeout=25, **kwargs)
            else:
                r = session.post(url, headers=HEADERS, timeout=25, **kwargs)
            r.raise_for_status()
            return r
        except Exception as e:
            logger.warning(f"[!] Request failed ({attempt}/{retries}) for {url}: {e}")
            if attempt < retries:
                sleep_time = backoff ** (attempt - 1)
                logger.info(f"[+] Retrying in {sleep_time}s...")
                time.sleep(sleep_time)
            else:
                logger.error(f"[-] Giving up on {url}")
                return None

def extract_android_packages_from_yeswehack(data):
    packages = set()

    def walk(obj):
        if isinstance(obj, dict):
            # Check for the common structure
            if "in_scope" in obj and isinstance(obj["in_scope"], list):
                for item in obj["in_scope"]:
                    if not isinstance(item, dict):
                        continue
                    t = item.get("target", "") or ""
                    typ = item.get("type", "") or ""
                    if typ == "mobile-application-android" and "play.google.com" in t:
                        m = re.search(r"[?&]id=([A-Za-z0-9._]+)", t)
                        if m:
                            packages.add(m.group(1))
            # Recurse other dict entries
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for i in obj:
                walk(i)

    walk(data)
    return sorted(packages)

def discover_uptodown_app_page(session, package, logger):
    search_url = "https://en.uptodown.com/android/search"
    logger.info(f"[+] Searching Uptodown for {package}")
    resp = safe_request(session, "post", search_url, logger, data={"q": package})
    if not resp:
        return None
    soup = BeautifulSoup(resp.text, "html.parser")

    selectors = [
        "div.item div.name a[href*='/android']",
        "div.item a[href*='/android']",
        "a[href*='.uptodown.com/android']",
        "a[href*='/android/']"
    ]
    for sel in selectors:
        a = soup.select_one(sel)
        if a and a.get("href"):
            href = a["href"].strip()
            # Normalize absolute/relative
            if href.startswith("http"):
                return href
            elif href.startswith("//"):
                return "https:" + href
            elif href.startswith("/"):
                return "https://en.uptodown.com" + href
            else:
                if href.startswith("http"):
                    return href
                return "https://" + href
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "/android" in href and "uptodown" in href:
            if href.startswith("http"):
                return href
            elif href.startswith("//"):
                return "https:" + href
            elif href.startswith("/"):
                return "https://en.uptodown.com" + href
            else:
                return "https://" + href
    logger.warning(f"[-] No uptodown search result found for {package}")
    return None


def download_from_app_page(session, app_page, outdir, logger):
    download_page = app_page.rstrip("/") + "/download"
    logger.info(f"[+] Fetching download page: {download_page}")
    resp = safe_request(session, "get", download_page, logger)
    if not resp:
        return None, None, None

    soup = BeautifulSoup(resp.text, "html.parser")
    file_type_row = soup.find("th", string=lambda s: s and s.strip().lower() == "file type")
    file_type = None
    if file_type_row:
        try:
            file_type = file_type_row.find_next_sibling("td").text.strip().lower()
        except Exception:
            file_type = None

    pkg_row = soup.find("th", string=lambda s: s and s.strip() == "Package Name")
    if pkg_row:
        try:
            pkg_name = pkg_row.find_next_sibling("td").text.strip()
        except Exception:
            pkg_name = None
    else:
        # fallback: infer from app_page host: https://bforbank.en.uptodown.com -> use host part 'bforbank' as fallback
        m = re.match(r"https?://([^.]+)\.", app_page)
        pkg_name = m.group(1) if m else f"unknown_{int(time.time())}"

    
    button = soup.select_one("button#detail-download-button[data-url]")
    if not button:
        logger.warning(f"[-] No download button found on {download_page}")
        return pkg_name, None, None
    data_url = button.get("data-url")
    if not data_url:
        logger.warning(f"[-] download button missing data-url on {download_page}")
        return pkg_name, None, None


    ext = "apk"
    if file_type and "xapk" in file_type:
        ext = "xapk"

    final_url = f"https://dw.uptodown.net/dwn/{data_url}/uptodown-{pkg_name}.{ext}"
    out_path = Path(outdir) / f"{pkg_name}.{ext}"
    if out_path.exists() and out_path.stat().st_size > 0:
        logger.info(f"[=] Already downloaded, skipping: {out_path}")
        return pkg_name, str(out_path), ext

    logger.info(f"[+] Downloading: {final_url} -> {out_path.name}")
    r = safe_request(session, "get", final_url, logger, stream=True)
    if not r:
        return pkg_name, None, ext

    total = int(r.headers.get("content-length", 0) or 0)
    try:
        with open(out_path, "wb") as f, tqdm(total=total, unit="B", unit_scale=True, desc=pkg_name, leave=False) as pbar:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    pbar.update(len(chunk))
        logger.info(f"[✓] Saved: {out_path}")
        return pkg_name, str(out_path), ext
    except Exception as e:
        logger.error(f"[!] Error saving {out_path}: {e}")
        try:
            if out_path.exists():
                out_path.unlink()
        except Exception:
            pass
        return pkg_name, None, ext

def extract_xapk(xapk_path, outdir, logger):
    import zipfile
    xp = Path(xapk_path)
    tempdir = Path(outdir) / f"{xp.stem}_xapk_extracted"
    tempdir.mkdir(parents=True, exist_ok=True)
    try:
        with zipfile.ZipFile(xapk_path, "r") as z:
            z.extractall(tempdir)
        # find first apk
        for p in tempdir.rglob("*.apk"):
            logger.info(f"[✓] Found inner APK inside XAPK -> {p}")
            return str(p), str(tempdir)
        logger.warning(f"[-] No APK found inside XAPK {xapk_path}")
        # nothing found
        return None, str(tempdir)
    except Exception as e:
        logger.error(f"[!] Failed to extract XAPK {xapk_path}: {e}")
        return None, None


def decompile_and_extract_manifest(apk_path, apktool_jar, logger):
    apk_path = Path(apk_path)
    decompiled_dir = Path(TEMP_DIR) / f"{apk_path.stem}_decompiled"
    # ensure clean
    if decompiled_dir.exists():
        shutil.rmtree(decompiled_dir, ignore_errors=True)
    decompiled_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"[+] Decompiling {apk_path}")
    try:
        result = subprocess.run(
            ["java", "-jar", str(apktool_jar), "d", "-f", str(apk_path), "-o", str(decompiled_dir)],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
        )
    except Exception as e:
        logger.error(f"[-] Apktool invocation error for {apk_path}: {e}")
        return None

    if result.returncode != 0:
        logger.warning(f"[-] Apktool failed for {apk_path}: {result.stderr.strip()[:400]}")
        return None

    manifest_src = decompiled_dir / "AndroidManifest.xml"
    if manifest_src.exists():
        Path(MANIFEST_DIR).mkdir(parents=True, exist_ok=True)
        manifest_dest = Path(MANIFEST_DIR) / f"{apk_path.stem}_AndroidManifest.xml"
        try:
            shutil.copy(manifest_src, manifest_dest)
            logger.info(f"[✓] Extracted AndroidManifest.xml -> {manifest_dest}")
            return str(manifest_dest)
        except Exception as e:
            logger.warning(f"[-] Failed copying manifest for {apk_path}: {e}")
            return None
    else:
        logger.warning(f"[-] AndroidManifest.xml not found in decompiled output for {apk_path}")
        return None

def download_worker(session, package, outdir, logger):
    info = {
        "package": package,
        "downloaded": False,
        "file_type": None,
        "apk_path": None,
        "manifest_path": None,
        "note": ""
    }
    app_page = discover_uptodown_app_page(session, package, logger)
    if not app_page:
        info["note"] = "No uptodown app page found"
        return info

    pkg_name, saved_path, ext = download_from_app_page(session, app_page, outdir, logger)
    if not saved_path:
        info["note"] = "Download failed"
        return info

    info["downloaded"] = True
    info["file_type"] = ext
    # if XAPK, extract inner apk, delete xapk (user requested)
    final_apk_path = saved_path
    temp_extract_dir = None
    if ext == "xapk":
        inner_apk, tempdir = extract_xapk(saved_path, outdir, logger)
        if not inner_apk:
            info["note"] = "XAPK extracted but no inner APK"
            # delete xapk
            try:
                Path(saved_path).unlink()
            except Exception:
                pass
            if tempdir:
                shutil.rmtree(tempdir, ignore_errors=True)
            return info
        # delete xapk file as requested
        try:
            Path(saved_path).unlink()
        except Exception:
            pass
        final_apk_path = inner_apk
        temp_extract_dir = tempdir

    info["apk_path"] = final_apk_path
    info["temp_extract_dir"] = temp_extract_dir  # for cleanup later
    return info

def decompile_worker(info, apktool_jar, logger):
    apk_path = info.get("apk_path")
    if not apk_path:
        return None
    manifest = decompile_and_extract_manifest(apk_path, apktool_jar, logger)
    if manifest:
        info["manifest_path"] = manifest
    else:
        info["note"] = info.get("note", "") + " | decompile_failed"
    return info

def create_excel_report(results, outdir):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    headers = ["Package", "Downloaded", "FileType", "APK Path", "Manifest Path", "Notes"]
    for col, head in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = head
        ws.cell(row=1, column=col).font = Font(bold=True)
    r = 2
    for info in results:
        ws.cell(row=r, column=1).value = info.get("package")
        ws.cell(row=r, column=2).value = "Yes" if info.get("downloaded") else "No"
        ws.cell(row=r, column=3).value = info.get("file_type") or ""
        ws.cell(row=r, column=4).value = info.get("apk_path") or ""
        ws.cell(row=r, column=5).value = info.get("manifest_path") or ""
        ws.cell(row=r, column=6).value = info.get("note") or ""
        r += 1
    excel_path = Path(outdir) / "download_summary.xlsx"
    wb.save(excel_path)
    return str(excel_path)

def zip_and_cleanup(outdir, log_path, logger):
    outdir = Path(outdir)
    # Zip APKs + log
    zip_apks = outdir / "apks_and_log.zip"
    with zipfile.ZipFile(zip_apks, "w", zipfile.ZIP_DEFLATED) as zf:
        # include apk files
        for apk in outdir.glob("*.apk"):
            zf.write(apk, arcname=apk.name)
        # include any extracted inner apk from temp dir (they will be moved into outdir for zipping)
        for inner in outdir.rglob("*_xapk_extracted/**/*.apk"):
            pass  # not needed, we already moved inner apk paths into outdir when extracting
        # include log
        if Path(log_path).exists():
            zf.write(log_path, arcname=Path(log_path).name)
    logger.info(f"[✓] APKs + log zipped -> {zip_apks}")

    zip_manifests = outdir / "manifests.zip"
    with zipfile.ZipFile(zip_manifests, "w", zipfile.ZIP_DEFLATED) as zf:
        for mf in Path(MANIFEST_DIR).glob("*.xml"):
            zf.write(mf, arcname=mf.name)
    logger.info(f"[✓] Manifests zipped -> {zip_manifests}")

    for apk in outdir.glob("*.apk"):
        try:
            apk.unlink()
        except Exception:
            pass
    for xapk_ex in outdir.glob("*_xapk_extracted"):
        try:
            shutil.rmtree(xapk_ex, ignore_errors=True)
        except Exception:
            pass
    for mf in Path(MANIFEST_DIR).glob("*.xml"):
        try:
            mf.unlink()
        except Exception:
            pass
    for folder in Path(TEMP_DIR).glob("*_decompiled"):
        try:
            shutil.rmtree(folder, ignore_errors=True)
        except Exception:
            pass
    logger.info("[✓] Cleanup complete (deleted APKs, XAPKs, manifests and decompiled folders).")

def main():
    parser = argparse.ArgumentParser(description="Uptodown multi-threaded downloader (APK/XAPK -> decompile manifests)")
    parser.add_argument("--json", required=True, help="Path to YesWeHack JSON")
    parser.add_argument("--outdir", default="downloads", help="Output directory")
    parser.add_argument("--threads", type=int, default=6, help="Number of threads for download/decompile")
    parser.add_argument("--skip-decompile", action="store_true", help="Skip decompilation stage")
    args = parser.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    # Logging
    log_file = outdir / "download_log.txt"
    logger = logging.getLogger("uptodown_downloader")
    logger.setLevel(logging.INFO)
    fh = logging.FileHandler(log_file, encoding="utf-8")
    ch = logging.StreamHandler(sys.stdout)
    fmt = logging.Formatter("[%(levelname)s] %(message)s")
    fh.setFormatter(fmt); ch.setFormatter(fmt)
    logger.addHandler(fh); logger.addHandler(ch)

    script_dir = Path(__file__).resolve().parent
    apktool = ensure_apktool(script_dir, logger)

    try:
        with open(args.json, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logger.error(f"[-] Failed to read JSON: {e}")
        return

    packages = extract_android_packages_from_yeswehack(data)
    logger.info(f"[✓] Extracted {len(packages)} package names from JSON")

    if not packages:
        logger.error("[-] No packages found in JSON. Exiting.")
        return

    session = requests.Session()

    logger.info(f"[+] Starting {len(packages)} downloads with {args.threads} threads")
    download_results = []
    with ThreadPoolExecutor(max_workers=args.threads) as ex:
        future_to_pkg = {ex.submit(download_worker, session, pkg, outdir, logger): pkg for pkg in packages}
        for fut in as_completed(future_to_pkg):
            pkg = future_to_pkg[fut]
            try:
                res = fut.result()
                download_results.append(res)
            except Exception as e:
                logger.error(f"[!] Exception downloading {pkg}: {e}")
                download_results.append({"package": pkg, "downloaded": False, "note": f"exception: {e}"})

    if not args.skip_decompile:
        logger.info(f"[+] Starting decompilation with {args.threads} threads")
        decompile_results = []
        with ThreadPoolExecutor(max_workers=args.threads) as ex:
            future_map = {}
            for info in download_results:
                if info.get("downloaded") and info.get("apk_path"):
                    future = ex.submit(decompile_worker, info, script_dir / APKTOOL_FILENAME, logger)
                    future_map[future] = info["package"]
                else:
                    decompile_results.append(info)  # not downloaded
            for fut in as_completed(future_map):
                pkg = future_map[fut]
                try:
                    res = fut.result()
                    if res:
                        decompile_results.append(res)
                except Exception as e:
                    logger.error(f"[!] Exception decompiling {pkg}: {e}")
                    # attach failure note to corresponding info
                    decompile_results.append({"package": pkg, "downloaded": True, "note": f"decompile exception: {e}"})
        info_map = {i["package"]: i for i in download_results}
        for d in decompile_results:
            info_map[d["package"]] = d
        final_results = list(info_map.values())
    else:
        final_results = download_results

    excel_path = create_excel_report(final_results, outdir)
    logger.info(f"[✓] Excel created -> {excel_path}")

    zip_and_cleanup(outdir, str(log_file), logger)
    logger.info("[✓] Process complete.")

if __name__ == "__main__":
    main()
